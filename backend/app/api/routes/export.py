from __future__ import annotations

import csv
import io
import uuid
from decimal import Decimal

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.api.deps import CurrentUser, DbSession, require_site_access
from app.models.financial_data import FinancialLineItem, FinancialStatement, StatementType
from app.models.fixed_assets import Asset
from app.models.hr import Employee, SalaryRecord
from app.models.intercompany import ICInvoice
from app.models.site import Site
from app.models.tax import TaxFiling
from app.models.user import UserRole
from app.services.consolidation import consolidate_financial_data, get_site_financial_data
from app.services.kpi import calculate_all_kpis

router = APIRouter(prefix="/export", tags=["export"])

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
NUMBER_FMT = '#,##0.00'


def _style_header_row(ws, col_count: int) -> None:
    """Apply header styling to the first row of a worksheet."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws) -> None:
    """Auto-fit column widths."""
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 3, 40)


def _xlsx_response(wb: Workbook, filename: str) -> StreamingResponse:
    """Convert workbook to streaming response."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _csv_response(rows: list[list[str]], headers: list[str], filename: str) -> StreamingResponse:
    """Build CSV streaming response."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _write_sheet_from_dicts(
    ws,
    headers: list[str],
    keys: list[str],
    items: list[dict],
) -> None:
    """Populate a worksheet from a list of dicts."""
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    for row_idx, item in enumerate(items, 2):
        for col_idx, key in enumerate(keys, 1):
            val = item.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if isinstance(val, (int, float, Decimal)):
                cell.number_format = NUMBER_FMT
    _style_header_row(ws, len(headers))
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Financial Statements Export
# ---------------------------------------------------------------------------


@router.get("/financial-statements/{site_id}")
async def export_financial_statements(
    site_id: uuid.UUID,
    db: DbSession,
    current_user: CurrentUser,
    period_year: int = Query(..., ge=2000, le=2100),
    period_month: int = Query(..., ge=1, le=12),
    statement_type: StatementType | None = Query(None),
    format: str = Query("xlsx", pattern="^(csv|xlsx)$"),
) -> StreamingResponse:
    """Export financial statements for a site as Excel or CSV."""
    await require_site_access(site_id, current_user)

    stmt = (
        select(FinancialStatement)
        .where(
            FinancialStatement.site_id == site_id,
            FinancialStatement.period_year == period_year,
            FinancialStatement.period_month == period_month,
        )
    )
    if statement_type is not None:
        stmt = stmt.where(FinancialStatement.statement_type == statement_type)

    result = await db.execute(stmt)
    statements = result.scalars().all()

    headers = ["Statement Type", "Line Item Code", "Line Item Name", "Amount", "Currency"]
    keys = ["statement_type", "line_item_code", "line_item_name", "amount", "currency"]
    items: list[dict] = []

    for s in statements:
        for li in s.line_items:
            items.append({
                "statement_type": s.statement_type.value,
                "line_item_code": li.line_item_code,
                "line_item_name": li.line_item_name,
                "amount": float(li.amount),
                "currency": s.currency,
            })

    fname = f"financial_statements_{site_id}_{period_year}_{period_month:02d}"

    if format == "csv":
        rows = [[str(item.get(k, "")) for k in keys] for item in items]
        return _csv_response(rows, headers, f"{fname}.csv")

    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Statements"
    _write_sheet_from_dicts(ws, headers, keys, items)
    return _xlsx_response(wb, f"{fname}.xlsx")


# ---------------------------------------------------------------------------
# KPIs Export
# ---------------------------------------------------------------------------


@router.get("/kpis/{site_id}")
async def export_kpis(
    site_id: uuid.UUID,
    db: DbSession,
    current_user: CurrentUser,
    period_year: int = Query(..., ge=2000, le=2100),
    period_month: int = Query(..., ge=1, le=12),
    format: str = Query("xlsx", pattern="^(csv|xlsx)$"),
) -> StreamingResponse:
    """Export KPIs for a site as Excel or CSV."""
    await require_site_access(site_id, current_user)

    data = await get_site_financial_data(db, site_id, period_year, period_month)
    merged: dict[str, Decimal] = {}
    for items_map in data.values():
        merged.update(items_map)
    kpis = calculate_all_kpis(merged)

    headers = ["Category", "KPI Name", "Value", "Unit", "Description"]
    keys = ["category", "name", "value", "unit", "description"]
    items: list[dict] = []

    for category, kpi_list in kpis.items():
        for kpi in kpi_list:
            items.append({
                "category": category.capitalize(),
                "name": kpi.name,
                "value": float(kpi.value) if kpi.value is not None else None,
                "unit": kpi.unit,
                "description": kpi.description or "",
            })

    fname = f"kpis_{site_id}_{period_year}_{period_month:02d}"

    if format == "csv":
        rows = [[str(item.get(k, "")) for k in keys] for item in items]
        return _csv_response(rows, headers, f"{fname}.csv")

    wb = Workbook()
    ws = wb.active
    ws.title = "KPIs"
    _write_sheet_from_dicts(ws, headers, keys, items)
    return _xlsx_response(wb, f"{fname}.xlsx")


# ---------------------------------------------------------------------------
# Employees Export
# ---------------------------------------------------------------------------


@router.get("/employees/{site_id}")
async def export_employees(
    site_id: uuid.UUID,
    db: DbSession,
    current_user: CurrentUser,
    format: str = Query("xlsx", pattern="^(csv|xlsx)$"),
) -> StreamingResponse:
    """Export employee list for a site."""
    await require_site_access(site_id, current_user)

    stmt = (
        select(Employee)
        .options(selectinload(Employee.position), selectinload(Employee.department))
        .where(Employee.site_id == site_id, Employee.is_active == True)  # noqa: E712
        .order_by(Employee.last_name, Employee.first_name)
    )
    result = await db.execute(stmt)
    employees = result.scalars().all()

    headers = ["Code", "First Name", "Last Name", "Email", "Department", "Position",
               "Type", "FTE", "Start Date", "Active"]
    keys = ["code", "first_name", "last_name", "email", "department", "position",
            "type", "fte", "start_date", "active"]
    items: list[dict] = []

    for e in employees:
        items.append({
            "code": e.employee_code,
            "first_name": e.first_name,
            "last_name": e.last_name,
            "email": e.email or "",
            "department": e.department.name if e.department else "",
            "position": e.position.title if e.position else "",
            "type": e.employment_type.value if e.employment_type else "",
            "fte": float(e.fte_ratio),
            "start_date": str(e.start_date) if e.start_date else "",
            "active": "Yes" if e.is_active else "No",
        })

    fname = f"employees_{site_id}"

    if format == "csv":
        rows = [[str(item.get(k, "")) for k in keys] for item in items]
        return _csv_response(rows, headers, f"{fname}.csv")

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    _write_sheet_from_dicts(ws, headers, keys, items)
    return _xlsx_response(wb, f"{fname}.xlsx")


# ---------------------------------------------------------------------------
# Salaries Export
# ---------------------------------------------------------------------------


@router.get("/salaries/{site_id}")
async def export_salaries(
    site_id: uuid.UUID,
    db: DbSession,
    current_user: CurrentUser,
    year: int = Query(..., ge=2000, le=2100),
    month: int = Query(..., ge=1, le=12),
    format: str = Query("xlsx", pattern="^(csv|xlsx)$"),
) -> StreamingResponse:
    """Export salary records for a site and period."""
    await require_site_access(site_id, current_user)

    stmt = (
        select(SalaryRecord)
        .join(Employee, SalaryRecord.employee_id == Employee.id)
        .options(selectinload(SalaryRecord.employee))
        .where(
            Employee.site_id == site_id,
            SalaryRecord.period_year == year,
            SalaryRecord.period_month == month,
        )
        .order_by(SalaryRecord.created_at)
    )
    result = await db.execute(stmt)
    records = result.scalars().all()

    headers = ["Employee Code", "Period", "Currency", "Gross Salary", "Net Salary",
               "Employer Taxes", "Employee Taxes", "Benefits", "Bonus", "Total Cost"]
    keys = ["code", "period", "currency", "gross", "net", "emp_tax", "ee_tax",
            "benefits", "bonus", "total"]
    items: list[dict] = []

    for r in records:
        items.append({
            "code": r.employee.employee_code if r.employee else "",
            "period": f"{r.period_year}-{r.period_month:02d}",
            "currency": r.currency,
            "gross": float(r.gross_salary),
            "net": float(r.net_salary),
            "emp_tax": float(r.employer_taxes),
            "ee_tax": float(r.employee_taxes),
            "benefits": float(r.benefits),
            "bonus": float(r.bonus),
            "total": float(r.total_cost),
        })

    fname = f"salaries_{site_id}_{year}_{month:02d}"

    if format == "csv":
        rows = [[str(item.get(k, "")) for k in keys] for item in items]
        return _csv_response(rows, headers, f"{fname}.csv")

    wb = Workbook()
    ws = wb.active
    ws.title = "Salaries"
    _write_sheet_from_dicts(ws, headers, keys, items)
    return _xlsx_response(wb, f"{fname}.xlsx")


# ---------------------------------------------------------------------------
# Assets Export
# ---------------------------------------------------------------------------


@router.get("/assets/{site_id}")
async def export_assets(
    site_id: uuid.UUID,
    db: DbSession,
    current_user: CurrentUser,
    format: str = Query("xlsx", pattern="^(csv|xlsx)$"),
) -> StreamingResponse:
    """Export asset register for a site."""
    await require_site_access(site_id, current_user)

    stmt = (
        select(Asset)
        .where(Asset.site_id == site_id)
        .order_by(Asset.asset_code)
    )
    result = await db.execute(stmt)
    assets = result.scalars().all()

    headers = ["Code", "Name", "Category", "Status", "Acquisition Date", "Currency",
               "Acquisition Cost", "Accum. Depreciation", "Net Book Value",
               "Useful Life (months)", "Method", "Location"]
    keys = ["code", "name", "category", "status", "acq_date", "currency",
            "cost", "accum_dep", "nbv", "life", "method", "location"]
    items: list[dict] = []

    for a in assets:
        items.append({
            "code": a.asset_code,
            "name": a.name,
            "category": a.category.value if a.category else "",
            "status": a.status.value if a.status else "",
            "acq_date": str(a.acquisition_date) if a.acquisition_date else "",
            "currency": a.currency,
            "cost": float(a.acquisition_cost),
            "accum_dep": float(a.accumulated_depreciation),
            "nbv": float(a.net_book_value),
            "life": a.useful_life_months,
            "method": a.depreciation_method.value if a.depreciation_method else "",
            "location": a.location or "",
        })

    fname = f"assets_{site_id}"

    if format == "csv":
        rows = [[str(item.get(k, "")) for k in keys] for item in items]
        return _csv_response(rows, headers, f"{fname}.csv")

    wb = Workbook()
    ws = wb.active
    ws.title = "Assets"
    _write_sheet_from_dicts(ws, headers, keys, items)
    return _xlsx_response(wb, f"{fname}.xlsx")


# ---------------------------------------------------------------------------
# IC Invoices Export
# ---------------------------------------------------------------------------


@router.get("/ic-invoices")
async def export_ic_invoices(
    db: DbSession,
    current_user: CurrentUser,
    site_id: uuid.UUID | None = Query(None),
    format: str = Query("xlsx", pattern="^(csv|xlsx)$"),
) -> StreamingResponse:
    """Export intercompany invoices."""
    stmt = select(ICInvoice).order_by(ICInvoice.invoice_date.desc())
    if site_id is not None:
        await require_site_access(site_id, current_user)
        stmt = stmt.where(
            (ICInvoice.sender_site_id == site_id) | (ICInvoice.receiver_site_id == site_id)
        )

    result = await db.execute(stmt)
    invoices = result.scalars().all()

    headers = ["Invoice No.", "Sender Site", "Receiver Site", "Date", "Due Date",
               "Currency", "Amount", "Category", "Status", "Description"]
    keys = ["number", "sender", "receiver", "date", "due", "currency", "amount",
            "category", "status", "description"]
    items: list[dict] = []

    for inv in invoices:
        items.append({
            "number": inv.invoice_number,
            "sender": str(inv.sender_site_id),
            "receiver": str(inv.receiver_site_id),
            "date": str(inv.invoice_date) if inv.invoice_date else "",
            "due": str(inv.due_date) if inv.due_date else "",
            "currency": inv.currency,
            "amount": float(inv.amount),
            "category": inv.category.value if inv.category else "",
            "status": inv.status.value if inv.status else "",
            "description": inv.description or "",
        })

    fname = "ic_invoices"

    if format == "csv":
        rows = [[str(item.get(k, "")) for k in keys] for item in items]
        return _csv_response(rows, headers, f"{fname}.csv")

    wb = Workbook()
    ws = wb.active
    ws.title = "IC Invoices"
    _write_sheet_from_dicts(ws, headers, keys, items)
    return _xlsx_response(wb, f"{fname}.xlsx")


# ---------------------------------------------------------------------------
# Tax Filings Export
# ---------------------------------------------------------------------------


@router.get("/tax-filings")
async def export_tax_filings(
    db: DbSession,
    current_user: CurrentUser,
    site_id: uuid.UUID | None = Query(None),
    year: int | None = Query(None, ge=2000, le=2100),
    format: str = Query("xlsx", pattern="^(csv|xlsx)$"),
) -> StreamingResponse:
    """Export tax filings."""
    stmt = select(TaxFiling).order_by(TaxFiling.due_date.desc())
    if site_id is not None:
        stmt = stmt.where(TaxFiling.site_id == site_id)
    if year is not None:
        stmt = stmt.where(TaxFiling.period_year == year)

    result = await db.execute(stmt)
    filings = result.scalars().all()

    headers = ["Site ID", "Filing Type", "Year", "Quarter", "Due Date",
               "Filed Date", "Status", "Amount", "Currency", "Notes"]
    keys = ["site_id", "type", "year", "quarter", "due", "filed",
            "status", "amount", "currency", "notes"]
    items: list[dict] = []

    for f in filings:
        items.append({
            "site_id": str(f.site_id),
            "type": f.filing_type.value if f.filing_type else "",
            "year": f.period_year,
            "quarter": f.period_quarter,
            "due": str(f.due_date) if f.due_date else "",
            "filed": str(f.filed_date) if f.filed_date else "",
            "status": f.status.value if f.status else "",
            "amount": float(f.amount) if f.amount else 0,
            "currency": f.currency or "",
            "notes": f.notes or "",
        })

    fname = "tax_filings"

    if format == "csv":
        rows = [[str(item.get(k, "")) for k in keys] for item in items]
        return _csv_response(rows, headers, f"{fname}.csv")

    wb = Workbook()
    ws = wb.active
    ws.title = "Tax Filings"
    _write_sheet_from_dicts(ws, headers, keys, items)
    return _xlsx_response(wb, f"{fname}.xlsx")


# ---------------------------------------------------------------------------
# Consolidated Report
# ---------------------------------------------------------------------------


@router.get("/consolidated-report")
async def export_consolidated_report(
    db: DbSession,
    current_user: CurrentUser,
    period_year: int = Query(..., ge=2000, le=2100),
    period_month: int = Query(..., ge=1, le=12),
    target_currency: str = Query("EUR", min_length=3, max_length=3),
    format: str = Query("xlsx", pattern="^(csv|xlsx)$"),
) -> StreamingResponse:
    """Full consolidated report with all statement types + KPIs in a multi-sheet workbook."""
    if current_user.role == UserRole.local_cfo:
        site_ids = [s.id for s in current_user.assigned_sites]
    else:
        result = await db.execute(select(Site.id).where(Site.is_active == True))  # noqa: E712
        site_ids = list(result.scalars().all())

    data = await consolidate_financial_data(
        db, site_ids, period_year, period_month, target_currency
    )

    # Compute KPIs
    merged: dict[str, Decimal] = {}
    for items_map in data.values():
        merged.update(items_map)
    kpis = calculate_all_kpis(merged)

    fname = f"consolidated_report_{period_year}_{period_month:02d}"

    if format == "csv":
        # For CSV, flatten everything into one table
        all_rows: list[list[str]] = []
        csv_headers = ["Sheet", "Line Item Code", "Amount"]
        for st_type, line_items in data.items():
            for code, amount in line_items.items():
                all_rows.append([st_type, code, str(amount)])
        # Add KPIs
        for category, kpi_list in kpis.items():
            for kpi in kpi_list:
                all_rows.append([f"KPI - {category}", kpi.name, str(kpi.value) if kpi.value is not None else ""])
        return _csv_response(all_rows, csv_headers, f"{fname}.csv")

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Sheet per statement type
    for st_type, line_items in data.items():
        ws = wb.create_sheet(title=st_type.replace("_", " ").title())
        ws_headers = ["Line Item Code", f"Amount ({target_currency})"]
        ws_keys = ["code", "amount"]
        items_list = [{"code": code, "amount": float(amount)} for code, amount in line_items.items()]
        _write_sheet_from_dicts(ws, ws_headers, ws_keys, items_list)

    # KPI sheet
    ws_kpi = wb.create_sheet(title="KPIs")
    kpi_headers = ["Category", "KPI Name", "Value", "Unit"]
    kpi_keys = ["category", "name", "value", "unit"]
    kpi_items: list[dict] = []
    for category, kpi_list in kpis.items():
        for kpi in kpi_list:
            kpi_items.append({
                "category": category.capitalize(),
                "name": kpi.name,
                "value": float(kpi.value) if kpi.value is not None else None,
                "unit": kpi.unit,
            })
    _write_sheet_from_dicts(ws_kpi, kpi_headers, kpi_keys, kpi_items)

    return _xlsx_response(wb, f"{fname}.xlsx")


__all__ = ["router"]
